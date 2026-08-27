"""Runs: build author rows, persist, diff, export, schedule."""
import collections
import json
import os
import plistlib
import subprocess
import sys
from datetime import datetime

sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
import census as X

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
RUNS = os.path.join(ROOT, "data", "runs")
EXPORTS = os.path.join(ROOT, "data", "exports")
PLIST_ID = "com.scifiniti.aautracker"
PLIST = os.path.expanduser("~/Library/LaunchAgents/%s.plist" % PLIST_ID)
for d in (RUNS, EXPORTS):
    os.makedirs(d, exist_ok=True)


def new_id():
    return datetime.now().strftime("%Y%m%d-%H%M%S")


# ------------------------------------------------------------ author rows
def _author_rows(log=print):
    """Rows carrying each author's name, AU-ID and PRINTED affiliation.

    Two sources, and the fallback is the whole reason a run works on a machine
    that has no census. Scopus will not give a co-author list to these keys --
    search view=COMPLETE and abstract view=FULL are both 401, and META carries
    only the first author -- so the author rows cannot be refetched from the
    API. Without a source here, build_slots returns nothing and every figure
    downstream (roster, colleges, per-author papers) comes out zero.

    1. the census export, when this machine has one
    2. engine/data/author_rows.json.gz -- a DERIVED file: names, AU-IDs, the
       printed affiliation and the correspondence line, with emails removed.
       No titles, abstracts, DOIs, ISSNs, funding or references.
    """
    import csv
    import gzip
    import json as _json
    csv.field_size_limit(10 ** 9)
    path = X.census_file("scopus_export.csv")
    if os.path.exists(path):
        with open(path, newline="", encoding="utf-8-sig") as fh:
            return list(csv.DictReader(fh))

    data = os.environ.get("AAU_DATA") or os.path.join(
        os.path.dirname(os.path.dirname(os.path.abspath(__file__))), "data")
    gz = os.path.join(data, "author_rows.json.gz")
    if not os.path.exists(gz):
        log("  no author rows available -- cannot build the census")
        return None
    with gzip.open(gz, "rt", encoding="utf-8") as fh:
        blob = _json.load(fh)
    log("  author rows from the shipped file: %d papers" % len(blob))
    return [{"EID": eid,
             "Author full names": "; ".join(
                 "%s (%s)" % (n, i) for n, i in zip(r["n"], r["i"])),
             "Authors with affiliations": r.get("a") or "",
             "Correspondence Address": r.get("c") or ""}
            for eid, r in blob.items()]


def build_slots(papers, log=print):
    """One row per author per paper, from the Scopus CSV export.

    The export is the only source that carries every author's own printed
    affiliation aligned with their AU-ID. OpenAlex truncates author lists at
    100 and often stores a shorter affiliation string, so it is not used here.
    """
    import csv
    import re
    csv.field_size_limit(10 ** 9)
    src = _author_rows(log)
    if src is None:
        return []

    named = re.compile(r"^(.*?)\s*\((\d{6,})\)\s*$")
    slots = []
    if True:
        for row in src:
            eid = (row.get("EID") or "").strip()
            if eid not in papers:
                continue
            p = papers[eid]
            names, ids = [], []
            for chunk in (row.get("Author full names") or "").split(";"):
                m = named.match(chunk.strip())
                if m:
                    names.append(m.group(1))
                    ids.append(m.group(2))
            affs = _split_affs(row.get("Authors with affiliations") or "")
            by_init = {}
            for anm, aff in affs:
                for k in X.initial_keys(anm):
                    by_init.setdefault(k, aff)
            corr = row.get("Correspondence Address") or ""
            corr_keys = set()
            for k in X.initial_keys(corr.split(";")[0][:80]):
                corr_keys.add(k)

            n = len(names)
            for i, nm in enumerate(names):
                aff = affs[i][1] if (len(affs) == n and i < len(affs)) else ""
                if not aff:
                    for k in X.initial_keys(nm):
                        if k in by_init:
                            aff = by_init[k]
                            break
                is_aau, sig = X.badge(
                    {"institutions": [], "raw_affiliation_strings": [aff] if aff else []},
                    p.get("doc_afids"))
                slots.append({
                    "eid": eid, "doi": p.get("doi"), "year": p.get("year"),
                    "journal": p.get("journal"), "doctype": p.get("doctype"),
                    "cited_by": p.get("cited_by"),
                    "author_name": nm, "scopus_auid": ids[i] if i < len(ids) else "",
                    "position": i + 1, "n_authors": n,
                    "is_first": i == 0,
                    "is_corresponding": bool(corr_keys & X.initial_keys(nm)),
                    "is_aau": is_aau, "sig_ror": sig.get("ror"),
                    "sig_string": sig.get("string"),
                    "raw_affiliation": aff,
                    "college": X.college_of([aff]) if is_aau else "",
                })
    log("  built %d author rows from the Scopus export" % len(slots))

    # The export covers the 1,330 papers of the original census and nothing
    # else, because co-author lists cannot be refetched: the Search API
    # accepts `field=author` and silently returns nothing, and every view that
    # carries them 401s with these keys. So a run over a wider window used to
    # collect 4,245 papers and then break them down by college using author
    # rows for barely a quarter of them -- a real total over a partial
    # breakdown, which reads as one consistent picture and is not.
    #
    # For the rest we have exactly one piece of authorship evidence, and it is
    # Scopus's own: AU-ID(x) returned this paper, so x is an author of it. That
    # yields faculty rows only -- students and outside co-authors stay unknown
    # on those papers, and `authors` is reported against the export as before.
    covered = {s["eid"] for s in slots}
    extra = 0
    for eid, p in papers.items():
        if eid in covered:
            continue
        for aid in (p.get("sweep_auids") or []):
            if not aid:
                continue
            slots.append({
                "eid": eid, "doi": p.get("doi"), "year": p.get("year"),
                "journal": p.get("journal"), "doctype": p.get("doctype"),
                "cited_by": p.get("cited_by"),
                "author_name": "", "scopus_auid": aid,
                "position": 0, "n_authors": 0,
                "is_first": False, "is_corresponding": False,
                "is_aau": True, "sig_ror": "", "sig_string": "",
                "raw_affiliation": "",
                "college": "",          # classify() fills it from the roster
                "from_sweep": True,
            })
            extra += 1
    if extra:
        log("  plus %d faculty rows from AU-ID provenance on %d papers the "
            "export does not cover"
            % (extra, len({s['eid'] for s in slots if s.get('from_sweep')})))
    return slots


def _split_affs(s):
    import re
    out = []
    for chunk in [x.strip() for x in (s or "").split(";") if x.strip()]:
        m = re.match(r"^(.+?,\s*[A-Z][\w'\-]*\.?(?:[\s\-][A-Z]\.?)*),\s*(.+)$", chunk)
        if m:
            out.append((m.group(1).strip(), m.group(2).strip()))
        else:
            parts = chunk.split(",", 1)
            out.append((parts[0].strip(), parts[1].strip() if len(parts) > 1 else ""))
    return out


def roster(slots):
    """Collapse AAU author rows into people."""
    agg = {}
    for s in slots:
        if not s.get("is_aau"):
            continue
        k = s.get("scopus_auid") or X.name_key(s.get("author_name") or "")
        if not k:
            continue
        r = agg.setdefault(k, {
            "key": k, "name": s.get("author_name"), "scopus_auid": s.get("scopus_auid", ""),
            "role": s.get("role"), "college": s.get("college"),
            "role_basis": s.get("role_basis"), "papers": 0, "corresponding": 0,
            "first_author": 0, "citations": 0, "years": set(), "eids": []})
        r["papers"] += 1
        r["corresponding"] += int(bool(s.get("is_corresponding")))
        r["first_author"] += int(bool(s.get("is_first")))
        r["citations"] += int(s.get("cited_by") or 0)
        if s.get("year"):
            r["years"].add(s["year"])
        r["eids"].append(s["eid"])
        # A sweep row carries an AU-ID but no printed name, so whichever
        # row lands first must not fix the person as nameless.
        if s.get("author_name") and not r.get("name"):
            r["name"] = s["author_name"]
        if s.get("role") and not r.get("role"):
            r["role"] = s["role"]
        if s.get("college") and not r.get("college"):
            r["college"] = s["college"]
    out = []
    for r in agg.values():
        r["years"] = sorted(r["years"])
        r["eids"] = sorted(set(r["eids"]))
        out.append(r)
    out.sort(key=lambda x: -x["papers"])
    return out


# ------------------------------------------------------------ persistence
def save(run_id, blob):
    d = os.path.join(RUNS, run_id)
    os.makedirs(d, exist_ok=True)
    blob["run_id"] = run_id
    blob["finished"] = datetime.now().isoformat(timespec="seconds")
    with open(os.path.join(d, "run.json"), "w") as fh:
        json.dump(blob, fh, default=str)
    with open(os.path.join(RUNS, "latest.txt"), "w") as fh:
        fh.write(run_id)
    return run_id


def load(run_id=None):
    run_id = run_id or latest_id()
    if not run_id:
        return {}
    p = os.path.join(RUNS, run_id, "run.json")
    if not os.path.exists(p):
        return {}
    with open(p) as fh:
        return json.load(fh)


def latest_id():
    p = os.path.join(RUNS, "latest.txt")
    if os.path.exists(p):
        return open(p).read().strip() or None
    ids = sorted(d for d in os.listdir(RUNS)
                 if os.path.isdir(os.path.join(RUNS, d)))
    return ids[-1] if ids else None


def recent(n=8):
    ids = sorted((d for d in os.listdir(RUNS)
                  if os.path.isdir(os.path.join(RUNS, d))), reverse=True)[:n]
    out = []
    for i in ids:
        b = load(i)
        if b:
            out.append({"run_id": i, "finished": b.get("finished"),
                        "papers": len(b.get("papers") or {}),
                        "people": len(b.get("roster") or []),
                        "faculty_version": b.get("faculty_version"),
                        "stats": b.get("stats", {})})
    return out


def results(run_id=None):
    b = load(run_id)
    if not b:
        return {"error": "no run yet"}
    rost = b.get("roster") or []
    by_col = collections.defaultdict(
        lambda: {"people": 0, "papers": set(), "citations": 0, "faculty": 0})
    for r in rost:
        c = r.get("college") or "Not stated"
        e = by_col[c]
        e["people"] += 1
        e["papers"].update(r.get("eids") or [])
        e["citations"] += r.get("citations") or 0
        e["faculty"] += int(r.get("role") == "Faculty")
    cols = [{"college": c, "people": v["people"], "papers": len(v["papers"]),
             "citations": v["citations"], "faculty": v["faculty"]}
            for c, v in by_col.items()]
    cols.sort(key=lambda x: -x["papers"])
    return {"run_id": b.get("run_id"), "finished": b.get("finished"),
            "stats": b.get("stats", {}), "faculty_version": b.get("faculty_version"),
            "colleges": cols, "roster": rost,
            "suggestions": b.get("suggestions") or []}


# ------------------------------------------------------------ delta
def delta(run_id, papers, rost):
    """What changed against the previous run.

    Keyed on EID, never on DOI or title: Scopus re-indexes, DOIs get added
    late, and titles get corrected. An EID is stable, so a paper that gains
    authors or shifts year is an UPDATE, not something new.
    """
    prev_id = latest_id()
    prev = load(prev_id) if prev_id and prev_id != run_id else {}
    if not prev:
        return {"first_run": True, "since": None, "new_papers": [],
                "new_people": [], "returning": [], "updated": []}

    old_papers = prev.get("papers") or {}
    old_roster = {r["key"]: r for r in (prev.get("roster") or [])}

    new_papers = [{"eid": e, "title": (p.get("title") or "")[:120],
                   "journal": p.get("journal"), "year": p.get("year")}
                  for e, p in papers.items() if e not in old_papers]

    updated = []
    for e, p in papers.items():
        o = old_papers.get(e)
        if not o:
            continue
        if (o.get("cited_by") or 0) != (p.get("cited_by") or 0):
            updated.append({"eid": e, "title": (p.get("title") or "")[:90],
                            "citations_was": o.get("cited_by"),
                            "citations_now": p.get("cited_by")})

    new_people, returning = [], []
    for r in rost:
        o = old_roster.get(r["key"])
        if not o:
            new_people.append({"name": r["name"], "college": r.get("college"),
                               "role": r.get("role"), "papers": r["papers"]})
        elif r["papers"] > o.get("papers", 0):
            returning.append({"name": r["name"], "college": r.get("college"),
                              "role": r.get("role"),
                              "was": o.get("papers", 0), "now": r["papers"]})

    return {"first_run": False, "since": prev_id,
            "new_papers": new_papers[:400],
            "new_people": new_people[:400],
            "returning": returning[:400],
            "updated": updated[:200],
            "counts": {"new_papers": len(new_papers), "new_people": len(new_people),
                       "returning": len(returning), "updated": len(updated)}}


# ------------------------------------------------------------ export
def export(run_id, kind="xlsx"):
    b = load(run_id)
    if not b:
        return {"ok": False, "error": "no run"}
    os.makedirs(EXPORTS, exist_ok=True)
    stamp = b.get("run_id")
    if kind == "xlsx":
        path = os.path.join(EXPORTS, "AAU_Tracker_%s.xlsx" % stamp)
        _xlsx(b, path)
    else:
        return {"ok": False, "error": "pptx export is wired after the first real run"}
    return {"ok": True, "path": path, "name": os.path.basename(path)}


def _xlsx(b, path):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter
    hdr = PatternFill("solid", fgColor="0A7A3A")   # AAU green
    hf = Font(color="FFFFFF", bold=True, size=10)

    def sheet(wb, title, headers, rows, widths):
        ws = wb.create_sheet(title[:31])
        ws.append(headers)
        for c in ws[1]:
            c.fill, c.font = hdr, hf
            c.alignment = Alignment(vertical="center", wrap_text=True)
        for r in rows:
            ws.append(r)
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = "A1:%s%d" % (get_column_letter(len(headers)),
                                          max(len(rows) + 1, 2))
        for i, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w
        return ws

    wb = Workbook()
    wb.remove(wb.active)
    res = results(b.get("run_id"))
    sheet(wb, "Summary", ["Metric", "Value"],
          [["Run", b.get("run_id")], ["Finished", b.get("finished")],
           ["Faculty list", b.get("faculty_version") or "(none)"],
           ["Papers", len(b.get("papers") or {})],
           ["People", len(res["roster"])],
           ["Faculty", sum(1 for r in res["roster"] if r.get("role") == "Faculty")],
           ["Student / external",
            sum(1 for r in res["roster"] if r.get("role") != "Faculty")]],
          [34, 40])
    sheet(wb, "People",
          ["Name", "Role", "College", "Papers", "First author", "Corresponding",
           "Citations", "Years", "Scopus AU-ID", "Basis"],
          [[r["name"], r.get("role"), r.get("college"), r["papers"],
            r["first_author"], r["corresponding"], r["citations"],
            "-".join(str(y) for y in (r.get("years") or [])[:1] +
                     (r.get("years") or [])[-1:]),
            r.get("scopus_auid"), r.get("role_basis")]
           for r in res["roster"]],
          [30, 20, 34, 9, 12, 13, 10, 12, 16, 46])
    sheet(wb, "By college",
          ["College", "People", "Faculty", "Papers", "Citations"],
          [[c["college"], c["people"], c["faculty"], c["papers"], c["citations"]]
           for c in res["colleges"]], [40, 10, 10, 10, 12])
    sug = b.get("suggestions") or []
    if sug:
        sheet(wb, "Suggested additions",
              ["Name", "College", "Papers", "Corresponding", "Why", "Scopus AU-ID"],
              [[s["name"], s.get("college"), s["papers"], s["corresponding"],
                s["why"], s.get("scopus_auid")] for s in sug],
              [30, 34, 9, 13, 34, 16])
    # The two sheets that make the workbook usable on its own: every paper
    # counted, and every paper thrown out with the reason. Without these you
    # can read the totals but cannot check them.
    papers = b.get("papers") or {}
    if papers:
        rows = []
        for eid, pp in papers.items():
            rows.append([eid, pp.get("doi") or "", pp.get("year") or "",
                         (pp.get("title") or "")[:300],
                         pp.get("journal") or "",
                         pp.get("cited_by") or 0,
                         pp.get("found_via") or "af_id",
                         pp.get("aau_evidence") or ""])
        rows.sort(key=lambda r: (-(r[2] or 0), r[3]))
        sheet(wb, "Papers",
              ["EID", "DOI", "Year", "Title", "Source", "Cited by",
               "Found via", "AAU evidence"],
              rows, [22, 26, 7, 70, 34, 9, 14, 26])

    rej = (b.get("stats") or {}).get("rejected_list") or b.get("rejected") or []
    if rej:
        sheet(wb, "Rejected papers",
              ["EID", "DOI", "Title", "Source", "Why it was rejected"],
              [[r.get("eid"), r.get("doi") or "", (r.get("title") or "")[:300],
                r.get("journal") or "", r.get("reason") or ""] for r in rej],
              [22, 26, 70, 34, 46])

    wb.save(path)


# ------------------------------------------------------------ schedule
def schedule_state():
    if not os.path.exists(PLIST):
        return {"enabled": False}
    try:
        with open(PLIST, "rb") as fh:
            p = plistlib.load(fh)
        cal = p.get("StartCalendarInterval") or {}
        return {"enabled": True, "weekday": cal.get("Weekday", 1),
                "hour": cal.get("Hour", 7)}
    except Exception:
        return {"enabled": False}


def set_schedule(enabled, weekday=1, hour=7):
    """launchd, not an in-app timer -- it must fire with the app closed."""
    if not enabled:
        subprocess.run(["launchctl", "unload", PLIST],
                       capture_output=True)
        if os.path.exists(PLIST):
            os.remove(PLIST)
        return {"enabled": False}
    os.makedirs(os.path.dirname(PLIST), exist_ok=True)
    plist = {
        "Label": PLIST_ID,
        "ProgramArguments": ["/usr/bin/python3",
                             os.path.join(ROOT, "app.py"), "--headless-run"],
        "WorkingDirectory": ROOT,
        "StartCalendarInterval": {"Weekday": int(weekday), "Hour": int(hour),
                                  "Minute": 0},
        "StandardOutPath": os.path.join(ROOT, "logs", "scheduled.log"),
        "StandardErrorPath": os.path.join(ROOT, "logs", "scheduled.err"),
        "RunAtLoad": False,
    }
    with open(PLIST, "wb") as fh:
        plistlib.dump(plist, fh)
    subprocess.run(["launchctl", "unload", PLIST], capture_output=True)
    r = subprocess.run(["launchctl", "load", PLIST], capture_output=True)
    return {"enabled": True, "weekday": weekday, "hour": hour,
            "loaded": r.returncode == 0,
            "error": (r.stderr or b"").decode()[:200] or None}
